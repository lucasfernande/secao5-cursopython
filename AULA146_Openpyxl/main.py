import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
planilhas = pedidos.sheetnames  # pegando o nome das páginas

planilha1 = pedidos['Página1']
print(planilha1['b4'].value)  # acessando o conteudo do campo da coluna B linha 4

print('#' * 40)

for campo in planilha1['b']:  # acessando toda a coluna B
    if campo.value is not None:
        print(campo.value)

print('#' * 40)

for linha in planilha1['a1:c2']:  # acessando o conteudo em determinado intervalo (a1:c2)
    for coluna in linha:
        print(coluna.value)

print('#' * 40)

for linha in planilha1:
    if linha[0].value is not None:
        print(linha[0].value, end=' ')
    if linha[1].value is not None:
        print(linha[1].value, end=' ')
    if linha[2].value is not None:
        print(linha[2].value)

print('#' * 40)

# modificando valores das planilhas

# planilha1['b3'].value = 2400.00 # não altera o valor da planilha original

# gerando novas linhas

"""
for linha in range(5, 16):
    numero_pedido = linha - 1
    # Coluna A = 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    planilha1.cell(linha, 3).value = round(uniform(10, 100), 2)

pedidos.save('planilha_modificada.xlsx')

"""
# criando uma planilha

nova_planilha = openpyxl.Workbook()
nova_planilha.create_sheet('Planilha1', 0)  # criando páginas
nova_planilha.create_sheet('Planilha2', 1)

planilha_criada1 = nova_planilha['Planilha1']
planilha_criada2 = nova_planilha['Planilha2']

for linha in range(1, 10):
    numero_pedido = linha - 1
    # Coluna A = 1
    planilha_criada1.cell(linha, 1).value = numero_pedido
    planilha_criada1.cell(linha, 2).value = 1200 + linha
    planilha_criada1.cell(linha, 3).value = round(uniform(10, 100), 2)


for linha in range(1, 10):
    planilha_criada2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 50), 2)}'
    planilha_criada2.cell(linha, 2).value = f'Maria {linha} {round(uniform(10, 50), 2)}'
    planilha_criada2.cell(linha, 3).value = f'Otávio {linha} {round(uniform(10, 50), 2)}'

nova_planilha.save('planilha_criada.xlsx')
